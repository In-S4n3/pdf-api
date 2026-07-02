"""Tests for the pdf_to_xlsx service (PDF tables -> editable .xlsx)."""

import io
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.api_errors import ApiError
from app.services import pdf_tools
from app.services.pdf_tools import _sheet_title, pdf_to_xlsx

sys.path.insert(0, str(Path(__file__).parent / "fixtures"))
import gen_excel_fixtures as g  # noqa: E402

ENCRYPTED_PDF = Path(__file__).parent / "fixtures" / "encrypted.pdf"


def _load(result: bytes):
    assert isinstance(result, bytes) and len(result) > 0
    return load_workbook(io.BytesIO(result))


# --- happy path -------------------------------------------------------------


def test_tables_pdf_returns_valid_xlsx():
    wb = _load(pdf_to_xlsx(g.tables_pdf()))
    assert len(wb.worksheets) == 1
    ws = wb.worksheets[0]
    assert ws["A1"].value == "Produto"
    assert ws["B1"].value == "Qtd"
    assert ws["C1"].value == "Preco"
    assert ws["A2"].value == "Caneta"
    assert ws["C3"].value == "4.00"


def test_three_tables_pdf_creates_three_sheets_in_page_order():
    wb = _load(pdf_to_xlsx(g.three_tables_pdf()))
    assert len(wb.worksheets) == 3
    titles = wb.sheetnames
    assert titles == ["Pag 1 Tabela 1", "Pag 1 Tabela 2", "Pag 2 Tabela 1"]
    # page 1 first table, page 2 last table -> content in page order
    assert wb["Pag 1 Tabela 1"]["A1"].value == "A1"
    assert wb["Pag 2 Tabela 1"]["A1"].value == "C1"


# --- error paths ------------------------------------------------------------


def test_prose_pdf_raises_no_tables_detected():
    with pytest.raises(ApiError) as exc:
        pdf_to_xlsx(g.prose_pdf())
    assert exc.value.status_code == 422
    assert exc.value.code == "no_tables_detected"


def test_scanned_pdf_raises_422():
    with pytest.raises(ApiError) as exc:
        pdf_to_xlsx(g.blank_pdf())
    assert exc.value.status_code == 422
    assert exc.value.code == "scanned_pdf"


def test_encrypted_pdf_raises_400():
    with pytest.raises(ApiError) as exc:
        pdf_to_xlsx(ENCRYPTED_PDF.read_bytes())
    assert exc.value.status_code == 400
    assert exc.value.code == "password_protected_pdf"


def test_corrupt_bytes_raise_400():
    with pytest.raises(ApiError) as exc:
        pdf_to_xlsx(b"this is not a pdf")
    assert exc.value.status_code == 400
    assert exc.value.code == "invalid_pdf"


def test_too_many_pages_raises_400(monkeypatch):
    monkeypatch.setattr(pdf_tools, "MAX_PAGES", 1)
    with pytest.raises(ApiError) as exc:
        pdf_to_xlsx(g.blank_pdf(pages=3))
    assert exc.value.status_code == 400
    assert exc.value.code == "too_many_pages"


def test_complexity_guard_raises_422(monkeypatch):
    monkeypatch.setattr(pdf_tools, "MAX_PATHS_PER_PAGE", 5)
    with pytest.raises(ApiError) as exc:
        pdf_to_xlsx(g.complex_paths_pdf(n_paths=30))
    assert exc.value.status_code == 422
    assert exc.value.code == "pdf_too_complex"


# --- regression guard: sparse legit table must NOT be pre-rejected ----------


def test_sparse_table_returns_200_not_scanned():
    wb = _load(pdf_to_xlsx(g.sparse_table_pdf()))
    assert len(wb.worksheets) == 1
    ws = wb.worksheets[0]
    assert ws["A1"].value == "1"
    assert ws["B2"].value == "4"


# --- formula/error injection neutralization ---------------------------------


def test_injection_cells_reload_as_literal_text():
    wb = _load(pdf_to_xlsx(g.injection_table_pdf()))
    ws = wb.worksheets[0]
    formula_cell = ws["A1"]
    error_cell = ws["B1"]
    assert formula_cell.data_type == "s"
    assert formula_cell.value == "=1+1"
    assert error_cell.data_type == "s"
    assert error_cell.value == "#REF!"
    # benign cells untouched
    assert ws["A2"].value == "normal"
    assert ws["B2"].value == "42"


# --- _sheet_title unit ------------------------------------------------------


def test_sheet_title_unique_forbidden_char_free_and_bounded():
    forbidden = set("*?:/[]\\")
    seen = set()
    for pno in range(1, 201):
        for ti in range(1, 21):
            title = _sheet_title(pno, ti)
            assert len(title) <= 31
            assert not (set(title) & forbidden)
            assert title not in seen  # (pno, ti) pairs are inherently unique
            seen.add(title)


# --- HTTP-level (TestClient, /v2 envelope) ----------------------------------

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def test_v2_pdf_to_excel_happy_returns_200(client):
    response = client.post(
        "/v2/pdf-to-excel",
        files={"file": ("test.pdf", io.BytesIO(g.tables_pdf()), "application/pdf")},
        data={"options": "{}"},
    )
    assert response.status_code == 200
    assert response.headers["content-type"] == XLSX_MIME
    assert response.headers["content-disposition"].endswith('.xlsx"')
    wb = load_workbook(io.BytesIO(response.content))
    assert wb.worksheets[0]["A1"].value == "Produto"


def test_v2_pdf_to_excel_prose_returns_422_envelope(client):
    response = client.post(
        "/v2/pdf-to-excel",
        files={"file": ("prose.pdf", io.BytesIO(g.prose_pdf()), "application/pdf")},
        data={"options": "{}"},
    )
    assert response.status_code == 422
    assert response.json()["error"]["code"] == "no_tables_detected"


def test_v2_pdf_to_excel_scanned_returns_422_envelope(client):
    response = client.post(
        "/v2/pdf-to-excel",
        files={"file": ("scan.pdf", io.BytesIO(g.blank_pdf()), "application/pdf")},
        data={"options": "{}"},
    )
    assert response.status_code == 422
    assert response.json()["error"]["code"] == "scanned_pdf"


def test_v2_pdf_to_excel_encrypted_returns_400_envelope(client):
    response = client.post(
        "/v2/pdf-to-excel",
        files={"file": ("enc.pdf", io.BytesIO(ENCRYPTED_PDF.read_bytes()), "application/pdf")},
        data={"options": "{}"},
    )
    assert response.status_code == 400
    assert response.json()["error"]["code"] == "password_protected_pdf"
